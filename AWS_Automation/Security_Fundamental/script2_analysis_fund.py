import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Define service categories
categories = {
    'Security and Identity': ['IAM', 'ACM', 'KMS', 'GuardDuty', 'Secret Manager', 'Secret Hub', 'SSM'],
    'Compute': ['Auto Scaling', 'EC2', 'ECS', 'EKS', 'Lambda', 'EMR', 'Step Functions'],
    'Storage': ['EBS', 'ECR', 'S3', 'DLM', 'Backup'],
    'Network': ['API Gateway', 'CloudFront', 'Route 53', 'VPC', 'ELB', 'ElasticCache', 'CloudTrail'],
    'Database': ['RDS', 'DynamoDB', 'Athena', 'Glue'],
    'Other': ['CloudFormation', 'CodeDeploy', 'Config', 'SNS', 'SQS', 'WorkSpaces', 'EventBridge', 'Config']
}

def create_simplified_report(report_file, final_report_file):
    # Read input report file (CSV or Excel)
    if report_file.endswith('.csv'):
        df = pd.read_csv(report_file)
    elif report_file.endswith('.xlsx'):
        df = pd.read_excel(report_file, engine='openpyxl')  # Specify openpyxl for .xlsx files
    elif report_file.endswith('.xls'):
        df = pd.read_excel(report_file, engine='xlrd')  # Specify xlrd for .xls files
    else:
        raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")

    # Add 'fixed' and 'feedback' columns (only for category sheets)
    df['fixed'] = ''  # Placeholder for checkbox (not interactive)
    df['feedback'] = ""  # Empty feedback field

    # Filter out rows based on 'status' column
    unsafe_df = df[df['status'] == 'alarm'].copy()
    safe_df = df[df['status'] != 'alarm'].copy()

    # Remove 'fixed' and 'feedback' columns from "safe" and "unsafe"
    safe_df = safe_df.drop(columns=['fixed', 'feedback'])
    unsafe_df = unsafe_df.drop(columns=['fixed', 'feedback'])

    # Initialize a dictionary to hold DataFrames for each category
    categorized_data = {category: pd.DataFrame() for category in categories}

    # Process each category and filter relevant services
    for category, services in categories.items():
        # Only rows with 'status' == 'alarm' are included in category sheets
        categorized_data[category] = unsafe_df[unsafe_df['title'].isin(services)]

    # Create a new Excel writer object using openpyxl engine
    with pd.ExcelWriter(final_report_file, engine='openpyxl') as writer:
        # Initialize openpyxl workbook and get the workbook and sheets
        workbook = writer.book

        # Define formats for the header and severity coloring
        header_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')
        critical_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')  # Purple for Critical
        high_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red for High
        medium_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange for Medium
        low_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for Low

        # Function to apply severity formatting using openpyxl
        def apply_severity_format(worksheet, df):
            if 'severity' in df.columns:
                severity_col = df.columns.get_loc('severity')  # Get column index of 'severity'
                for row_num, severity in enumerate(df['severity'], start=2):  # Start from row 2 (skip header)
                    cell = worksheet.cell(row=row_num, column=severity_col + 1)
                    if severity.lower() == 'critical':
                        cell.fill = critical_fill
                    elif severity.lower() == 'high':
                        cell.fill = high_fill
                    elif severity.lower() == 'medium':
                        cell.fill = medium_fill
                    elif severity.lower() == 'low':
                        cell.fill = low_fill

        # Write "safe" and "unsafe" DataFrames to separate sheets
        for sheet_name, df_data in [('safe', safe_df), ('unsafe', unsafe_df)]:
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for col_num, value in enumerate(df_data.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.value = value
                cell.fill = header_fill  # Apply header formatting
            apply_severity_format(worksheet, df_data)

        # Write each category DataFrame to a separate sheet
        for category, data in categorized_data.items():
            if not data.empty:
                # Write to Excel
                data.to_excel(writer, sheet_name=category, index=False)
                worksheet = writer.sheets[category]
                for col_num, value in enumerate(data.columns.values):
                    cell = worksheet.cell(row=1, column=col_num + 1)
                    cell.value = value
                    cell.fill = header_fill  # Apply header formatting
                apply_severity_format(worksheet, data)

        # Create a summary sheet
        summary_data = pd.concat([
            unsafe_df[unsafe_df['severity'] == 'critical'],
            unsafe_df[unsafe_df['severity'] == 'high'],
            unsafe_df[unsafe_df['severity'] == 'medium'],
            unsafe_df[unsafe_df['severity'] == 'low']
        ])
        summary_columns = ['title', 'control_description', 'reason', 'status', 'severity', 'Recommendation Steps/Approach']
        summary_data = summary_data[summary_columns]
        summary_data.to_excel(writer, sheet_name="Summary", index=False)
        summary_ws = writer.sheets["Summary"]
        for col_num, value in enumerate(summary_data.columns.values):
            cell = summary_ws.cell(row=1, column=col_num + 1)
            cell.value = value
            cell.fill = header_fill  # Apply header formatting
        apply_severity_format(summary_ws, summary_data)

    print(f"Final simplified report saved as {final_report_file}")


def main():
    # Ask the user to input the report file name
    report_file = input("Enter the report file name (e.g., aws_compliance_benchmark_all_controls_benchmark_vested_with_priorities.csv): ").strip()

    # Get the input file's base name (without path) and extension
    base_name = os.path.splitext(os.path.basename(report_file))[0]

    # Create a unique file name using a timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_file_name = f"{base_name}_simplified_report_{timestamp}.xlsx"

    # Set the output path to the reports directory
    reports_directory = os.path.dirname(os.path.abspath(__file__))  # Get the current script's directory
    final_report_file = os.path.join(reports_directory, unique_file_name)

    create_simplified_report(report_file, final_report_file)


if __name__ == "__main__":
    main()

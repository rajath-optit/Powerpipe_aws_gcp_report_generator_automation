import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import re

# Define color fills for Excel
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Purple for Critical severity

# Load the input file and the database
def load_data(input_file, recommendation_file):
    if input_file.endswith(".xlsx"):
        df_input = pd.read_excel(input_file)
    elif input_file.endswith(".csv"):
        df_input = pd.read_csv(input_file, low_memory=False)
    else:
        raise ValueError("Unsupported file type")

    # Load recommendation database
    df_recommendation = pd.read_excel(recommendation_file)
    
    return df_input, df_recommendation

# Function to clean control_title by removing leading numbers and spaces
def clean_control_title(control_title):
    # This will remove leading numbers and spaces (e.g., "13 CloudFront distributions..." becomes "CloudFront distributions...")
    return re.sub(r'^\d+\s+', '', control_title)

# Update severity and add recommendations
def update_severity_and_recommendation(df_input, df_recommendation):
    for idx, row in df_input.iterrows():
        control_title = row["control_title"]

        # Clean the control_title to remove any leading numbers and spaces
        cleaned_control_title = clean_control_title(control_title)

        # Search for the cleaned control_title in the recommendation database
        matching_row = df_recommendation[df_recommendation["control_title"] == cleaned_control_title]

        if not matching_row.empty:
            recommendation = matching_row.iloc[0]["Recommendation Steps/Approach"]

            # Update the recommendation
            df_input.at[idx, "Recommendation Steps/Approach"] = recommendation
        else:
            # If no match is found, set a default recommendation
            df_input.at[idx, "Recommendation Steps/Approach"] = "No recommendation available"

    return df_input

# Write the output file with timestamped name
def write_output(df_input, output_file):
    # Temporarily include severity_color for formatting
    temp_columns = [
        "title", "control_title", "control_description", "region", "account_id", 
        "resource", "reason", "severity", "Recommendation Steps/Approach", "status"
    ]
    df_input = df_input[temp_columns]

    # Save the updated data frame to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_input.to_excel(writer, index=False, sheet_name="Sheet1")

        # Apply the color formatting to the severity column based on its value
        wb = writer.book
        ws = wb["Sheet1"]
        for idx, row in df_input.iterrows():
            severity = row["severity"]
            cell = ws.cell(row=idx+2, column=temp_columns.index("severity")+1)
            if severity.lower() == "high":
                cell.fill = red_fill
            elif severity.lower() == "medium":
                cell.fill = orange_fill
            elif severity.lower() == "low":
                cell.fill = yellow_fill
            elif severity.lower() == "critical":
                cell.fill = purple_fill  # Color for Critical severity

    wb.save(output_file)

def main():
    # Get file names from user
    input_file = input("Enter the input file name (CSV or Excel): ")
    recommendation_file = "PowerPipeControls_Annotations.xlsx"  # Recommendation file containing control titles and recommendations
    
    # Add timestamp to output file name for uniqueness
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = input_file.split(".")[0] + f"_updated_{timestamp}.xlsx"  # Save with timestamp

    # Load data
    df_input, df_recommendation = load_data(input_file, recommendation_file)

    # Update input with recommendations
    updated_df = update_severity_and_recommendation(df_input, df_recommendation)

    # Write output file with severity color formatting
    write_output(updated_df, output_file)
    print(f"Updated file saved as {output_file}")

if __name__ == "__main__":
    main()

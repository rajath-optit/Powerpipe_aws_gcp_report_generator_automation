import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
from datetime import datetime

# Define color fills for Excel
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Load the input file and the database
def load_data(input_file, priority_file):
    if input_file.endswith(".xlsx"):
        df_input = pd.read_excel(input_file)
    elif input_file.endswith(".csv"):
        df_input = pd.read_csv(input_file, low_memory=False)
    else:
        raise ValueError("Unsupported file type")

    # Load priority database
    df_priority = pd.read_excel(priority_file)
    
    return df_input, df_priority

# Match control_title and update with priority and recommendations
def update_priority_and_recommendation(df_input, df_priority):
    for idx, row in df_input.iterrows():
        control_title = row["control_title"]
        status = row["status"]

        # Search for the control_title in the priority database
        matching_row = df_priority[df_priority["control_title"] == control_title]

        if not matching_row.empty:
            priority = matching_row.iloc[0]["priority"]
            recommendation = matching_row.iloc[0]["Recommendation Steps/Approach"]

            # If the status is ok, info, or skip, color the priority green and set text
            if status in ["ok", "info", "skip"]:
                df_input.at[idx, "priority"] = "Safe/Well Architected"
                df_input.at[idx, "Recommendation Steps/Approach"] = recommendation
                df_input.at[idx, "priority_color"] = "00FF00"  # Green color
            else:
                # For other statuses, assign the appropriate priority and recommendation
                df_input.at[idx, "priority"] = priority
                df_input.at[idx, "Recommendation Steps/Approach"] = recommendation
                
                # Color the priority column based on the priority
                if priority == "High":
                    df_input.at[idx, "priority_color"] = "FF0000"  # Red
                elif priority == "Medium":
                    df_input.at[idx, "priority_color"] = "FFA500"  # Orange
                elif priority == "Low":
                    df_input.at[idx, "priority_color"] = "FFFF00"  # Yellow
        else:
            # If no match is found, set default values
            df_input.at[idx, "priority"] = "No data"
            df_input.at[idx, "Recommendation Steps/Approach"] = "No recommendation available"
            df_input.at[idx, "priority_color"] = "FFFFFF"  # White or default

        # Add new columns for feedback and fixed checkbox
        df_input.at[idx, "Feedback"] = ""  # Placeholder for feedback
        
        # Set 'Fixed' to False only if the status is 'alarm'
        if status == "alarm":
            df_input.at[idx, "Fixed"] = False  # Placeholder for 'Fixed' checkbox
        else:
            df_input.at[idx, "Fixed"] = True  # Set to True for other statuses

    return df_input

# Write the output file with timestamped name
def write_output(df_input, output_file):
    # Temporarily include priority_color for formatting
    temp_columns = [
        "title", "control_title", "control_description", "region", "account_id", 
        "resource", "reason", "priority", "Recommendation Steps/Approach", "status", "priority_color", 
        "Feedback", "Fixed"
    ]
    df_input = df_input[temp_columns]

    # Save the updated data frame to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_input.drop(columns=["priority_color"], inplace=False).to_excel(writer, index=False, sheet_name="All Data")  # Exclude priority_color in the final sheet

        # Apply the color formatting to the priority column based on the 'priority_color' field
        wb = writer.book
        ws = wb["All Data"]
        for idx, row in df_input.iterrows():
            priority_color = row["priority_color"]
            if priority_color != "FFFFFF":  # Skip if no color is needed
                cell = ws.cell(row=idx+2, column=temp_columns.index("priority")+1)
                if priority_color == "FF0000":
                    cell.fill = red_fill
                elif priority_color == "FFA500":
                    cell.fill = orange_fill
                elif priority_color == "FFFF00":
                    cell.fill = yellow_fill
                elif priority_color == "00FF00":
                    cell.fill = green_fill

        # Separate data into compliance and non-compliance sheets
        df_compliance = df_input[df_input['status'].isin(['ok', 'info', 'skip'])]
        df_non_compliance = df_input[df_input['status'] == 'alarm']

        # Write compliance and non-compliance data to separate sheets
        df_compliance.drop(columns=["priority_color"], inplace=True)
        df_non_compliance.drop(columns=["priority_color"], inplace=True)

        df_compliance.to_excel(writer, index=False, sheet_name="Compliance")
        df_non_compliance.to_excel(writer, index=False, sheet_name="Non-Compliance")

        # Create Pivot tables for compliance and non-compliance using pandas
        pivot_compliance = pd.pivot_table(df_compliance, index='control_title', columns='status', aggfunc='size', fill_value=0)
        pivot_non_compliance = pd.pivot_table(df_non_compliance, index='control_title', columns='status', aggfunc='size', fill_value=0)

        # Write pivot tables to Excel
        pivot_compliance.to_excel(writer, sheet_name="Compliance Pivot", startrow=1, header=True, index=True)
        pivot_non_compliance.to_excel(writer, sheet_name="Non-Compliance Pivot", startrow=1, header=True, index=True)

        # Access the sheets where the pivot tables are written
        ws_compliance = wb["Compliance"]
        ws_non_compliance = wb["Non-Compliance"]

        # Create a line chart for the compliance vs. non-compliance count
        chart_compliance = LineChart()
        chart_compliance.title = "Compliance Issues"
        chart_compliance.style = 13  # Use style 13 for a line chart
        chart_compliance.x_axis.title = 'Control Title'
        chart_compliance.y_axis.title = 'Count'

        chart_non_compliance = LineChart()
        chart_non_compliance.title = "Non-Compliance Issues"
        chart_non_compliance.style = 13
        chart_non_compliance.x_axis.title = 'Control Title'
        chart_non_compliance.y_axis.title = 'Count'

        # Reference for compliance chart
        data_ref_compliance = Reference(ws_compliance, min_col=1, min_row=1, max_col=2, max_row=len(df_compliance) + 1)
        chart_compliance.add_data(data_ref_compliance, titles_from_data=True)

        # Reference for non-compliance chart
        data_ref_non_compliance = Reference(ws_non_compliance, min_col=1, min_row=1, max_col=2, max_row=len(df_non_compliance) + 1)
        chart_non_compliance.add_data(data_ref_non_compliance, titles_from_data=True)

        # Add charts to the sheets
        ws_compliance.add_chart(chart_compliance, "F5")  # Position the chart at F5
        ws_non_compliance.add_chart(chart_non_compliance, "F5")  # Position the chart at F5

    wb.save(output_file)

def main():
    # Get file names from user
    input_file = input("Enter the input file name (CSV or Excel): ")
    priority_file = "PowerPipeControls_Annotations.xlsx"  # Database file containing control titles, priorities, etc.
    
    # Add timestamp to output file name for uniqueness
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = input_file.split(".")[0] + f"_updated_{timestamp}.xlsx"  # Save with timestamp

    # Load data
    df_input, df_priority = load_data(input_file, priority_file)

    # Update input with priority and recommendations
    updated_df = update_priority_and_recommendation(df_input, df_priority)

    # Write output file with priority color formatting and separate sheets
    write_output(updated_df, output_file)
    print(f"Updated file saved as {output_file}")

if __name__ == "__main__":
    main()
